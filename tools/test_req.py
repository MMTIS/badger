#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
reg_test.py

Execute each line of a given input file as a command (with parameters) and
record the command, its return code, and the last 5000 characters of the
combined output (stdout+stderr) into an Excel file (.xlsx). Cell values are
sanitized to remove illegal characters so no exceptions are raised by openpyxl.

Usage:
  python reg_test.py commands.txt results.xlsx
  python reg_test.py commands.txt results.xlsx --timeout 60
  python reg_test.py commands.txt results.xlsx --no-shell

Notes:
- Lines starting with '#' or that are blank are ignored.
- By default, commands are executed through the system shell (shell=True).
- Use --no-shell to run commands directly without a shell.
"""

import argparse
import os
import sys
import shlex
import subprocess
from typing import List, Tuple, Optional

# Excel writing uses openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError as e:
    print("Error: openpyxl is required to write .xlsx files.\n" "Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

# Constants
EXCEL_CELL_MAX_CHARS = 32767
OUTPUT_TAIL_CHARS = 5000  # keep only last 5000 characters of output


def read_commands(path: str, encoding: Optional[str] = None) -> List[str]:
    """
    Read lines from a file and return the list of command lines,
    skipping blank lines and lines starting with '#'.
    """
    encodings_to_try = [encoding] if encoding else ["utf-8-sig", "utf-8", "latin-1"]
    last_err = None
    for enc in encodings_to_try:
        try:
            with open(path, "r", encoding=enc) as f:
                lines = f.readlines()
            break
        except Exception as err:  # try next encoding
            last_err = err
    else:
        raise last_err  # re-raise if no encoding worked

    commands = []
    for raw in lines:
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        commands.append(line)
    return commands


def run_command(command: str, shell: bool, timeout: Optional[float], text_encoding: str) -> Tuple[int, str]:
    """
    Execute a single command line.
    Returns a tuple: (return_code, combined_output)
    """
    try:
        if shell:
            completed = subprocess.run(
                command,
                shell=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,  # combine stderr into stdout
                timeout=timeout,
                text=True,
                encoding=text_encoding,
                errors="replace",
            )
        else:
            args = shlex.split(command, posix=(os.name != "nt"))
            completed = subprocess.run(
                args,
                shell=False,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                timeout=timeout,
                text=True,
                encoding=text_encoding,
                errors="replace",
            )
        rc = completed.returncode
        output = completed.stdout if completed.stdout is not None else ""
        return rc, output
    except subprocess.TimeoutExpired as ex:
        # Collect any partial output and mark timeout
        output = ""
        if ex.stdout:
            output += ex.stdout
        if ex.stderr:
            output += ex.stderr
        output += f"\n[TimeoutExpired: command exceeded {timeout} seconds]"
        return 124, output  # 124 is a common timeout code
    except FileNotFoundError as ex:
        return 127, f"[FileNotFoundError] {ex}"
    except Exception as ex:
        return 1, f"[Error running command] {type(ex).__name__}: {ex}"


# ------------- Excel-safe text helpers ----------------

# Try to use openpyxl's illegal characters regex if available; otherwise fallback
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE as OPENPYXL_ILLEGAL_RE  # openpyxl <=3.1
except Exception:
    OPENPYXL_ILLEGAL_RE = None

import re

# Fallback regex: strip ASCII control chars except tab(0x09), LF(0x0A), CR(0x0D)
FALLBACK_ILLEGAL_RE = re.compile(r'[\x00-\x08\x0B\x0C\x0E-\x1F]')


def remove_illegal_chars(s: str) -> str:
    """
    Remove characters that Excel/openpyxl cannot handle in cell text.
    Keeps \t, \n, \r and removes other ASCII control characters.
    """
    if not isinstance(s, str):
        s = str(s)
    # Normalize newlines
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    try:
        if OPENPYXL_ILLEGAL_RE:
            return OPENPYXL_ILLEGAL_RE.sub("", s)
    except Exception:
        # If importing regex from openpyxl fails at runtime for any reason
        pass
    return FALLBACK_ILLEGAL_RE.sub("", s)


def excel_tail_text(s: str, tail_len: int) -> str:
    """
    Sanitize for Excel and return only the last 'tail_len' characters.
    Ensures result does not exceed Excel's max cell length.
    """
    s = remove_illegal_chars(s)
    # Last tail_len characters
    if tail_len is not None and tail_len >= 0 and len(s) > tail_len:
        s = s[-tail_len:]
    # Excel absolute max safeguard (tail_len is 5000 < 32767, but keep this for safety)
    if len(s) > EXCEL_CELL_MAX_CHARS:
        s = s[-EXCEL_CELL_MAX_CHARS:]
    return s


def excel_safe_text(s: str) -> str:
    """
    Sanitize for Excel without tailing, but still enforce the maximum cell length.
    """
    s = remove_illegal_chars(s)
    if len(s) > EXCEL_CELL_MAX_CHARS:
        s = s[:EXCEL_CELL_MAX_CHARS]
    return s


# ------------------------------------------------------


def init_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    headers = ["Script (with parameters)", "Return code", "Output (last 5000 chars)"]
    ws.append(headers)
    # Styling headers
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
    # Column widths (approximate)
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 100
    return wb


def append_result(ws, command: str, return_code: int, output: str) -> None:
    """
    Append a single sanitized result row to the worksheet.
    Output column contains only the last 5000 characters.
    """
    # Sanitize command and output to avoid illegal characters
    safe_cmd = excel_safe_text(command)
    safe_out = excel_tail_text(output, OUTPUT_TAIL_CHARS)

    # Append row and set wrapping
    ws.append([safe_cmd, return_code, safe_out])
    row_idx = ws.max_row
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=row_idx, column=1).alignment = wrap
    ws.cell(row=row_idx, column=3).alignment = wrap


def ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    if parent and not os.path.exists(parent):
        os.makedirs(parent, exist_ok=True)


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Execute commands from a file and collect results in an Excel workbook (output column stores only the last 5000 characters)."
    )
    parser.add_argument("input_file", help="Path to the text file containing commands (one per line).")
    parser.add_argument("output_excel", help="Path to the output Excel file (.xlsx).")
    parser.add_argument("--timeout", type=float, default=None, help="Per-command timeout in seconds (default: no timeout).")
    parser.add_argument("--shell", dest="shell", action="store_true", help="Execute commands through the system shell (default).")
    parser.add_argument("--no-shell", dest="shell", action="store_false", help="Execute commands directly without a shell.")
    parser.set_defaults(shell=True)

    parser.add_argument("--encoding", default="utf-8", help="Text encoding for decoding process output (default: utf-8).")

    return parser.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)

    out_ext = os.path.splitext(args.output_excel)[1].lower()
    if out_ext not in (".xlsx",):
        print(f"Warning: Output file does not have .xlsx extension: {args.output_excel}", file=sys.stderr)

    try:
        commands = read_commands(args.input_file)
    except Exception as e:
        print(f"Failed to read input file '{args.input_file}': {e}", file=sys.stderr)
        return 2

    if not commands:
        print("No commands to execute (file empty or only comments).", file=sys.stderr)

    wb = init_workbook()
    ws = wb.active

    total = len(commands)
    for idx, cmd in enumerate(commands, start=1):
        print(f"\n\n=============================================================\n[{idx}/{total}] Running: {cmd}")
        rc, out = run_command(cmd, shell=args.shell, timeout=args.timeout, text_encoding=args.encoding)
        # Append sanitized results (output limited to last 5000 chars)
        print(f'Returned {rc}.\n')
        print(f"\n\n-------------------------------------------------------------\n\n")
        print(out)
        try:
            append_result(ws, cmd, rc, out)
        except Exception as e:
            # As a last resort, coerce via encoding replacement to avoid any surprises
            fallback_cmd = excel_safe_text(str(cmd).encode("utf-8", errors="replace").decode("utf-8", errors="replace"))
            fallback_out = excel_tail_text(str(out).encode("utf-8", errors="replace").decode("utf-8", errors="replace"), OUTPUT_TAIL_CHARS)
            ws.append([fallback_cmd, rc, fallback_out])

    # Save workbook
    try:
        ensure_parent_dir(args.output_excel)
        wb.save(args.output_excel)
        print(f"Results written to: {args.output_excel}")
    except Exception as e:
        print(f"Failed to write Excel file '{args.output_excel}': {e}", file=sys.stderr)
        return 3

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
